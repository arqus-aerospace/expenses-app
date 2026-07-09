"""Generate the expense-tracker.xlsx template embedded in the web app.

Sheets:
  - Data:      the "Expenses" table (created here so it ships pre-seeded with
               the imported June/July 2026 expenses; Graph appends to it)
  - Dashboard: live formulas (month total, averages, YTD, pending count),
               a rolling 12-month totals block, a per-category block, and
               native Excel charts wired to those blocks.

Writes tools/expense-tracker.xlsx (preview) and js/xlsx-template.js (embedded
base64). Keep COLUMNS in js/config.js in sync with HEADERS below.
"""
import base64
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

#            A     B            C       D           E        F         G           H              I        J          K      L      M           N          O              P         Q            R
HEADERS = ["ID", "Submitted", "Date", "Employee", "Email", "Vendor", "Category", "Description", "Gross", "VAT %", "VAT", "Net", "Currency", "Payment", "ReceiptFile", "Status", "DecidedBy", "DecidedOn"]
WIDTHS = [10, 18, 12, 18, 26, 30, 18, 40, 11, 8, 10, 11, 10, 22, 30, 12, 16, 18]

CATEGORIES = ["Travel Expenses", "Hardware", "Software/SaaS", "Infrastructure",
              "Office & Team", "Marketing/Sales", "Legal & Notary", "Miscellaneous"]

INK = "0B0B0B"
HEAD_FILL = "1F3A5F"      # deep blue header
ACCENT = "2A78D6"

EPOCH = datetime(1899, 12, 30)
def serial(y, m, d):
    return (datetime(y, m, d) - EPOCH).days

# Imported expenses (June/July 2026). "Ready for Accountant" -> Approved,
# no status yet -> Pending (approve in the app). VAT for the HIT receipt was
# blank in the source; inferred at 19% (beverages).
# (id, date(y,m,d), vendor, category, description, gross, vat_rate, payment, receipt_on_file, status)
CCC, BT = "Company Credit Card", "Bank Transfer"
SEED = [
    ("RE-001", (2026, 6, 1),  "Engel und Hain GbR", "Legal & Notary", "Pre-Seed legal advisory", 616.64, 0.19, BT, True, "Approved"),
    ("RE-002", (2026, 6, 1),  "Engel und Hain GbR", "Legal & Notary", "Notary cost estimation", 195.61, 0.19, BT, True, "Approved"),
    ("RE-003", (2026, 6, 20), "Gerstäcker Dresden GmbH & Co. KG", "Marketing/Sales", "Paper for business cards", 27.60, 0.19, CCC, True, "Approved"),
    ("RE-004", (2026, 6, 22), "Daniel Franz Copy Planet Dresden", "Marketing/Sales", "Printing of business cards", 80.00, 0.19, CCC, True, "Approved"),
    ("RE-005", (2026, 6, 25), "bavAIRia e.V.", "Miscellaneous", "Membership fee 2026", 225.00, 0.0, BT, False, "Approved"),
    ("RE-006", (2026, 6, 28), "Amazon", "Infrastructure", "Workbench & Storage", 187.96, 0.19, CCC, True, "Pending"),
    ("RE-007", (2026, 6, 29), "Elegoo", "Hardware", "PAHT-CF filament", 72.98, 0.19, CCC, True, "Pending"),
    ("RE-008", (2026, 6, 29), "Bambulab", "Hardware", "PLA & nozzles", 175.83, 0.19, CCC, True, "Pending"),
    ("RE-009", (2026, 6, 29), "Vevor", "Hardware", "Fume extractor soldering", 104.90, 0.19, CCC, True, "Pending"),
    ("RE-010", (2026, 6, 30), "Anthropic", "Software/SaaS", "", 75.11, 0.19, CCC, True, "Pending"),
    ("RE-011", (2026, 7, 1),  "HIT", "Office & Team", "Drinks company dinner", 12.61, 0.19, CCC, True, "Approved"),
    ("RE-012", (2026, 7, 1),  "Lidl", "Office & Team", "Dinner company dinner", 45.72, 0.19, CCC, True, "Approved"),
    ("RE-013", (2026, 7, 6),  "Deutsche Bahn", "Travel Expenses", "Travel Berlin Defence Tech Week", 199.00, 0.07, CCC, True, "Approved"),
    ("RE-014", (2026, 7, 6),  "Deutsche Bahn", "Travel Expenses", "Travel Berlin Defence Tech Week", 67.49, 0.07, CCC, True, "Approved"),
    ("RE-015", (2026, 7, 6),  "Booking.com", "Travel Expenses", "Travel Berlin Defence Tech Week", 144.00, 0.0, CCC, False, "Pending"),
    ("RE-016", (2026, 7, 6),  "Counter UAS Forum", "Travel Expenses", "Berlin Defence Tech Week", 250.00, 0.0, CCC, False, "Pending"),
    ("RE-017", (2026, 7, 6),  "Reimbursement Flight Employment interview", "Travel Expenses", "Cyril Austria flights", 285.00, 0.07, BT, True, "Approved"),
    ("RE-018", (2026, 7, 6),  "Max Emanuel Brauerei", "Office & Team", "Dinner minus 150 EUR Voucher", 20.00, 0.19, CCC, True, "Approved"),
]

wb = Workbook()

# ---------------- Data sheet ----------------
ws = wb.active
ws.title = "Data"
head_font = Font(bold=True, color="FFFFFF", size=11)
head_fill = PatternFill("solid", fgColor=HEAD_FILL)
for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = head_font
    c.fill = head_fill
    c.alignment = Alignment(vertical="center")
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

# Number formats for the first 2000 data rows. The app writes dates as Excel
# date serials so the Dashboard SUMIFS formulas compare numerically.
for r in range(2, 2002):
    ws.cell(row=r, column=2).number_format = 'yyyy-mm-dd hh:mm'  # Submitted
    ws.cell(row=r, column=3).number_format = 'yyyy-mm-dd'        # Date
    for col in (9, 11, 12):                                      # Gross, VAT, Net
        ws.cell(row=r, column=col).number_format = '#,##0.00'
    ws.cell(row=r, column=10).number_format = '0%'               # VAT %
    ws.cell(row=r, column=18).number_format = 'yyyy-mm-dd hh:mm' # DecidedOn

# Seed rows
for i, (rid, (y, m, d), vendor, cat, desc, gross, rate, pay, onfile, status) in enumerate(SEED):
    r = 2 + i
    vat = round(gross - gross / (1 + rate), 2)
    row = [rid, serial(y, m, d), serial(y, m, d), "", "", vendor, cat, desc,
           gross, rate, vat, round(gross - vat, 2), "EUR", pay,
           "on file" if onfile else "", status, "", ""]
    for c, v in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=v)

# The Excel table the app appends to (header + seeded rows).
last_col = get_column_letter(len(HEADERS))
table = Table(displayName="Expenses", ref=f"A1:{last_col}{1 + len(SEED)}")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(table)

# Status conditional formatting
ok = PatternFill("solid", fgColor="D9F2D9")
warn = PatternFill("solid", fgColor="FFF3CC")
bad = PatternFill("solid", fgColor="F8D7D7")
rng = "P2:P2001"
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Approved"'], fill=ok))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pending"'], fill=warn))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Rejected"'], fill=bad))

# ---------------- Dashboard sheet ----------------
db = wb.create_sheet("Dashboard")
db.sheet_view.showGridLines = False
db.column_dimensions["A"].width = 2
for col, w in zip("BCDEFGH", [26, 16, 3, 26, 16, 3, 16]):
    db.column_dimensions[col].width = w

title_font = Font(bold=True, size=16, color=INK)
kpi_label = Font(size=10, color="52514E")
kpi_value = Font(bold=True, size=14, color=INK)
sect_font = Font(bold=True, size=11, color=INK)
thin = Side(style="thin", color="E1E0D9")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

db["B2"] = "Expense dashboard"
db["B2"].font = title_font
db["B3"] = "Live view — recalculates whenever the app adds or approves an expense."
db["B3"].font = Font(size=9, color="898781")

# Data columns: Date = C, Gross = I, VAT = K, Category = G, Status = P.
# "Counted" spend = everything that is not Rejected.
NOT_REJ = 'Data!$P:$P,"<>Rejected"'

kpis = [
    ("This month (gross)", f'=SUMIFS(Data!$I:$I,Data!$C:$C,">="&EOMONTH(TODAY(),-1)+1,Data!$C:$C,"<="&EOMONTH(TODAY(),0),{NOT_REJ})'),
    ("Last month", f'=SUMIFS(Data!$I:$I,Data!$C:$C,">="&EOMONTH(TODAY(),-2)+1,Data!$C:$C,"<="&EOMONTH(TODAY(),-1),{NOT_REJ})'),
    ("Average / month (12m)", f'=SUMIFS(Data!$I:$I,Data!$C:$C,">="&EOMONTH(TODAY(),-12)+1,{NOT_REJ})/12'),
    ("Year to date", f'=SUMIFS(Data!$I:$I,Data!$C:$C,">="&DATE(YEAR(TODAY()),1,1),{NOT_REJ})'),
    ("VAT year to date", f'=SUMIFS(Data!$K:$K,Data!$C:$C,">="&DATE(YEAR(TODAY()),1,1),{NOT_REJ})'),
    ("Awaiting approval", '=COUNTIF(Data!$P:$P,"Pending")'),
]
row = 5
for label, formula in kpis:
    db.cell(row=row, column=2, value=label).font = kpi_label
    v = db.cell(row=row, column=3, value=formula)
    v.font = kpi_value
    v.number_format = '#,##0.00' if "COUNTIF" not in formula else '0'
    v.alignment = Alignment(horizontal="right")
    db.cell(row=row, column=2).border = box
    v.border = box
    row += 1

# Rolling 12 months block (K:M helper area, charted)
db["K1"] = "MonthStart"; db["L1"] = "Month"; db["M1"] = "Total"
for c in ("K1", "L1", "M1"):
    db[c].font = sect_font
for i in range(12):
    r = 2 + i
    off = 11 - i
    db[f"K{r}"] = f"=EOMONTH(TODAY(),-{off + 1})+1"
    db[f"K{r}"].number_format = "yyyy-mm-dd"
    db[f"L{r}"] = f'=TEXT(K{r},"mmm yy")'
    db[f"M{r}"] = (f'=SUMIFS(Data!$I:$I,Data!$C:$C,">="&K{r},'
                   f'Data!$C:$C,"<="&EOMONTH(K{r},0),{NOT_REJ})')
    db[f"M{r}"].number_format = '#,##0.00'

# Category block (O:P helper area, charted)
db["O1"] = "Category"; db["P1"] = "Total (12m)"
db["O1"].font = sect_font; db["P1"].font = sect_font
for i, cat in enumerate(CATEGORIES):
    r = 2 + i
    db[f"O{r}"] = cat
    db[f"P{r}"] = (f'=SUMIFS(Data!$I:$I,Data!$G:$G,O{r},'
                   f'Data!$C:$C,">="&EOMONTH(TODAY(),-12)+1,{NOT_REJ})')
    db[f"P{r}"].number_format = '#,##0.00'

# Line chart — expenses over time (last 12 months)
line = LineChart()
line.title = "Expenses per month (last 12 months)"
line.style = 12
line.height = 8
line.width = 18
line.y_axis.numFmt = '#,##0'
line.y_axis.majorGridlines = None
data = Reference(db, min_col=13, min_row=1, max_row=13)   # M
cats = Reference(db, min_col=12, min_row=2, max_row=13)   # L
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
s = line.series[0]
s.smooth = False
s.graphicalProperties.line.solidFill = ACCENT
s.graphicalProperties.line.width = 20000  # 2pt in EMU-ish units
line.legend = None
db.add_chart(line, "B13")

# Bar chart — spend by category
bar = BarChart()
bar.type = "bar"
bar.title = "Spend by category (last 12 months)"
bar.style = 12
bar.height = 8
bar.width = 18
data = Reference(db, min_col=16, min_row=1, max_row=1 + len(CATEGORIES))  # P
cats = Reference(db, min_col=15, min_row=2, max_row=1 + len(CATEGORIES))  # O
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.series[0].graphicalProperties.solidFill = ACCENT
bar.legend = None
bar.y_axis.majorGridlines = None
db.add_chart(bar, "B30")

wb.calculation.fullCalcOnLoad = True

repo = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
out = os.path.join(repo, "tools", "expense-tracker.xlsx")
wb.save(out)

with open(out, "rb") as f:
    b64 = base64.b64encode(f.read()).decode()

js = ("// Auto-generated by tools/make_template.py — do NOT edit by hand.\n"
      "// Base64 of the expense-tracker.xlsx template (Data sheet with the\n"
      "// pre-seeded \"Expenses\" table + live Dashboard sheet with formulas and\n"
      "// native Excel charts). Uploaded to SharePoint by the app on first run\n"
      "// if the workbook does not exist yet.\n"
      "export const XLSX_TEMPLATE_B64 =\n")
lines = [f'  "{b64[i:i + 100]}"' for i in range(0, len(b64), 100)]
js += " +\n".join(lines) + ";\n"

with open(os.path.join(repo, "js", "xlsx-template.js"), "w") as f:
    f.write(js)

print("wrote", out, "and js/xlsx-template.js —", len(b64) * 3 // 4, "xlsx bytes,",
      len(SEED), "seed rows")
